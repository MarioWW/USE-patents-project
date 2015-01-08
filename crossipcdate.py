import openpyxl

class crossipcdate:
	def __init__(self, mainFileName):#, rawinput):
		self.mainFileName = mainFileName
		self.wb = openpyxl.load_workbook(self.mainFileName)
		self.sheet = self.wb.get_active_sheet()

	def Parse_ipcdate(self,ExportName):
		ipccodes = {}
		data = (self.sheet.columns[12], self.sheet.columns[10])
		for i in range(0, len(data[0])):
		    for entry in data[1][i].value.split(";"):
		        entry = entry.strip()
		        ipccode = ""
		        ipccode = entry[:4]
		        try:
		            datelist = ipccodes[ipccode]
		        except:
		            datelist = {}
		        for code in data[0][i].value.split(";"):
		            code = code.strip()
		            yearcode = code[-4:]
		            monthDict={'Jan':'01', 'Feb':'02', 'Mar':'03', 'Apr':'04', 'May':'05', 'Jun':'06', 'Jul':'07', 'Aug':'08', 'Sep':'09', 'Oct':'10', 'Nov':'11', 'Dec':'12'}
		            #month
		            month=code[-8:-5]
		            month = monthDict.get(month)
		            datecode = str(yearcode) + "-" + month
		            try:
		                datelist[datecode] += 1
		            except:
		                datelist[datecode] = 1
		        ipccodes[ipccode] = datelist
		exportwb = openpyxl.Workbook()
		sheet = exportwb.get_active_sheet()
		columns = []

		for datelist in list(ipccodes.values()):
		    for date in list(datelist.keys()):
		        if date not in columns:
		            columns.append(date)

		columns = sorted(columns)
		for i in range(0,len(columns)):
		    sheet.cell(row=0,column=(i+1)).value = columns[i]

		for i in range(0, len(list(ipccodes.keys()))):
		    sheet.cell(row=(i+1),column=0).value = list(ipccodes.keys())[i]
		    for date in list(list(ipccodes.values())[i].keys()):
		        sheet.cell(row=(i+1),column=(columns.index(date) + 1)).value = list(ipccodes.values())[i][date]
		exportwb.save(ExportName)
